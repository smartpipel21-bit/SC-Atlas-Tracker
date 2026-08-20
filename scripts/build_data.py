#!/usr/bin/env python3
"""
Build data/technologies.json from data/Technologies_source.xlsx.

Run this whenever the source spreadsheet is updated:
    python3 scripts/build_data.py

It normalizes the messy free-text "Development Stage" column into a small
set of stage buckets (used for filtering and badge coloring), and parses
the concentration column into a numeric-or-null field plus a bucket label.
"""
import json
import re
import sys
from pathlib import Path
from datetime import datetime

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SOURCE_XLSX = ROOT / "data" / "Technologies_source.xlsx"
OUTPUT_JSON = ROOT / "data" / "technologies.json"

# Ordered from earliest to most mature. Order matters: it drives the
# ordinal color ramp in the dashboard (light -> dark = early -> mature).
STAGE_BUCKETS = [
    "Research / Academic",
    "Preclinical",
    "Platform / Feasibility",
    "Pre-registration / Late-stage",
    "Approved / Marketed",
]
STAGE_OTHER = "Varies by Program"

# Substring rules, checked in order, against the raw stage text (lowercased).
# First match wins. Keep this list in sync with the buckets above.
STAGE_RULES = [
    ("varies by partner program", STAGE_OTHER),
    ("approved", "Approved / Marketed"),
    ("commercially available", "Approved / Marketed"),
    ("commercial (", "Approved / Marketed"),
    ("pre-registration", "Pre-registration / Late-stage"),
    ("platform", "Platform / Feasibility"),
    ("research", "Research / Academic"),
    ("academic", "Research / Academic"),
    ("preclinical", "Preclinical"),
]

CONCENTRATION_BUCKETS = [
    (0, 300, "< 300 mg/mL"),
    (300, 450, "300–450 mg/mL"),
    (450, 600, "450–600 mg/mL"),
    (600, float("inf"), "600+ mg/mL"),
]
CONCENTRATION_NOT_DISCLOSED = "Not disclosed"


def classify_stage(raw: str) -> str:
    if not raw:
        return STAGE_OTHER
    low = raw.lower()
    for needle, bucket in STAGE_RULES:
        if needle in low:
            return bucket
    return STAGE_OTHER


def classify_concentration(value) -> str:
    if not isinstance(value, (int, float)):
        return CONCENTRATION_NOT_DISCLOSED
    for lo, hi, label in CONCENTRATION_BUCKETS:
        if lo <= value < hi:
            return label
    return CONCENTRATION_NOT_DISCLOSED


def to_iso(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    return None


def main():
    if not SOURCE_XLSX.exists():
        print(f"Source file not found: {SOURCE_XLSX}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() for h in rows[0]]

    records = []
    for raw_row in rows[1:]:
        if raw_row is None or all(v is None for v in raw_row):
            continue
        d = dict(zip(header, raw_row))

        raw_stage = (d.get("Development Stage") or "").strip()
        raw_conc = d.get("Concentration Achieved (numeric mg/mL)")
        conc_numeric = raw_conc if isinstance(raw_conc, (int, float)) else None

        record = {
            "id": d.get("Technology ID"),
            "name": d.get("Technology Name"),
            "company": d.get("Owner / Company"),
            "stage_raw": raw_stage,
            "stage_bucket": classify_stage(raw_stage),
            "type": d.get("Technology Type(s)"),
            "concentration_text": d.get("Concentration Achieved (text)"),
            "concentration_numeric": conc_numeric,
            "concentration_bucket": classify_concentration(conc_numeric),
            "needle_size": d.get("Needle Size"),
            "mechanism": d.get("Mechanism Summary"),
            "date_added": to_iso(d.get("Date Added")),
            "last_reviewed": to_iso(d.get("Last Reviewed")),
        }
        records.append(record)

    # Sort newest-reviewed first for the "recently reviewed" panel.
    records.sort(key=lambda r: r["last_reviewed"] or "", reverse=True)

    stage_order = STAGE_BUCKETS + [STAGE_OTHER]
    # Types ordered by frequency (most common first) so the most prevalent
    # approach gets the first (most legible) categorical color slot. This
    # order is fixed once computed here -- the dashboard must not re-sort
    # by whatever subset a filter leaves visible (color follows the entity,
    # never its rank in the current view).
    type_counts = {}
    for r in records:
        if r["type"]:
            type_counts[r["type"]] = type_counts.get(r["type"], 0) + 1
    types = sorted(type_counts, key=lambda t: (-type_counts[t], t))
    companies = sorted({r["company"] for r in records if r["company"]})
    concentration_order = [b[2] for b in CONCENTRATION_BUCKETS] + [CONCENTRATION_NOT_DISCLOSED]

    last_reviewed_dates = [r["last_reviewed"] for r in records if r["last_reviewed"]]

    payload = {
        "generated_from": SOURCE_XLSX.name,
        "last_updated": max(last_reviewed_dates) if last_reviewed_dates else None,
        "stage_order": stage_order,
        "type_order": types,
        "company_order": companies,
        "concentration_order": concentration_order,
        "technologies": records,
    }

    OUTPUT_JSON.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n")
    print(f"Wrote {len(records)} technologies to {OUTPUT_JSON}")


if __name__ == "__main__":
    main()
